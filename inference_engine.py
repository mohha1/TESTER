from __future__ import annotations

import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import cv2
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from openvino.runtime import Core
except Exception:
    from openvino import Core


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


@dataclass
class ModelPackage:
    model_xml: Path
    info_yml: Path
    meta_json: Path | None


@dataclass
class InferenceSettings:
    mode: str = "auto"
    device: str = "CPU"
    save_artifacts: bool = True
    save_only_ng_artifacts: bool = False

    connectivity: int = 8
    min_blob_area: int = 0
    color_order: str = "RGB"
    scale_01: bool = True
    normalize_mode: str = "never"
    threshold_op: str = ">="
    score_metric: str = "max_blob_area"


@dataclass
class InferenceResult:
    file: str
    product: str
    anomaly_score: float
    blob_count: int
    max_blob_area: int
    sum_blob_area: int
    ng_pixels_total: int
    candidate_path: str
    binary_path: str
    masked_path: str
    overlay_path: str
    width: int
    height: int


def discover_model_package(path: str | os.PathLike[str]) -> ModelPackage:
    base = Path(path).expanduser().resolve()
    if base.is_file() and base.name.lower() == "model.xml":
        model_xml = base
    elif (base / "model.xml").is_file():
        model_xml = base / "model.xml"
    else:
        candidates = sorted(
            base.rglob("model.xml"),
            key=lambda p: (0 if "openvino" in [x.lower() for x in p.parts] else 1, len(p.parts)),
        )
        if not candidates:
            raise FileNotFoundError(f"No model.xml found under: {base}")
        model_xml = candidates[0]

    search_roots = [model_xml.parent, *model_xml.parents]
    info_yml = None
    for root in search_roots:
        matches = sorted(list(root.glob("info*.yml")) + list(root.glob("info*.yaml")))
        if matches:
            info_yml = matches[0]
            break
    if info_yml is None:
        raise FileNotFoundError("No info.yml/info_*.yml found near the model package.")

    meta = model_xml.parent / "meta_data.json"
    return ModelPackage(model_xml=model_xml, info_yml=info_yml, meta_json=meta if meta.is_file() else None)


def imread_unicode_safe(path: str):
    try:
        data = np.fromfile(path, dtype=np.uint8)
        img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
        if img is None:
            return None
        if len(img.shape) == 2:
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        elif img.shape[2] == 4:
            img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
        return img
    except Exception:
        return None


def load_info_yml(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_meta_json(path: Path | None) -> dict | None:
    if not path or not path.is_file():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def normalize_anomaly_map(raw_map: np.ndarray, meta: dict) -> np.ndarray:
    vmin = float(meta["min"])
    vmax = float(meta["max"])
    if vmax - vmin < 1e-12:
        return np.zeros_like(raw_map, dtype=np.float32)
    return np.clip((raw_map - vmin) / (vmax - vmin), 0.0, 1.0).astype(np.float32)


def preprocess_to_model(img_bgr: np.ndarray, input_shape, input_dtype, settings: InferenceSettings):
    _, c, h, w = list(input_shape)
    resized = cv2.resize(img_bgr, (w, h), interpolation=cv2.INTER_AREA)
    image = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB) if settings.color_order.upper() == "RGB" else resized
    x = image.astype(np.float32) if np.dtype(input_dtype).kind == "f" else image.astype(input_dtype)
    if np.dtype(input_dtype).kind == "f" and settings.scale_01:
        x = x / 255.0
    x = np.transpose(x, (2, 0, 1))[None, ...]
    if c == 1 and x.shape[1] == 3:
        gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY).astype(np.float32)
        if settings.scale_01:
            gray /= 255.0
        x = gray[None, None, :, :]
    return x


def pick_anomaly_map_output(compiled_model, inference_result) -> np.ndarray:
    outputs = compiled_model.outputs
    if len(outputs) == 1:
        return np.array(inference_result[outputs[0]])
    best = None
    best_ndim = -1
    for out in outputs:
        arr = np.array(inference_result[out])
        if arr.ndim > best_ndim:
            best = arr
            best_ndim = arr.ndim
    return best


def squeeze_to_hw(output_array: np.ndarray) -> np.ndarray:
    m = np.array(output_array)
    if m.ndim == 4:
        m = m[0, 0]
    elif m.ndim == 3:
        m = m[0]
    elif m.ndim != 2:
        raise RuntimeError(f"Unexpected output shape: {m.shape}")
    return m.astype(np.float32)


def to_uint8_map_like_sample(raw_map: np.ndarray) -> np.ndarray:
    return np.clip(raw_map * 255.0 + 0.5, 0.0, 255.0).astype(np.uint8)


def apply_binary_threshold(cand8: np.ndarray, binary_thr: int, settings: InferenceSettings) -> np.ndarray:
    if settings.threshold_op == ">":
        return (cand8 > binary_thr).astype(np.uint8)
    return (cand8 >= binary_thr).astype(np.uint8)


def remove_small_blobs(mask01: np.ndarray, min_area: int, connectivity: int):
    if min_area <= 0:
        return mask01
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(mask01, connectivity=connectivity)
    cleaned = np.zeros_like(mask01, dtype=np.uint8)
    for lbl in range(1, num_labels):
        if stats[lbl, cv2.CC_STAT_AREA] >= min_area:
            cleaned[labels == lbl] = 1
    return cleaned


def compute_blob_areas(mask01: np.ndarray, connectivity: int):
    num_labels, _, stats, _ = cv2.connectedComponentsWithStats(mask01.astype(np.uint8), connectivity=connectivity)
    if num_labels <= 1:
        return 0, [], 0
    areas = stats[1:, cv2.CC_STAT_AREA].astype(int).tolist()
    return len(areas), areas, int(max(areas)) if areas else 0


def compute_anomaly_score(cand8, max_blob_area, sum_blob_area, settings: InferenceSettings):
    if settings.score_metric == "sum_blob_area":
        return sum_blob_area
    if settings.score_metric == "max_pixel":
        return int(cand8.max())
    return max_blob_area


def build_visuals(img_bgr: np.ndarray, candidate8: np.ndarray, mask01: np.ndarray):
    h, w = candidate8.shape
    original_resized = cv2.resize(img_bgr, (w, h), interpolation=cv2.INTER_AREA)
    heat = cv2.applyColorMap(candidate8, cv2.COLORMAP_JET)
    masked = np.zeros_like(heat)
    masked[mask01.astype(bool)] = heat[mask01.astype(bool)]
    overlay = cv2.addWeighted(original_resized, 0.65, heat, 0.35, 0)
    contours, _ = cv2.findContours((mask01 * 255).astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(overlay, contours, -1, (255, 255, 255), 1)
    return masked, overlay


def _write_image(path: Path, image: np.ndarray) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(str(path), image)
    return str(path)


def run_batch(
    model_path: str | os.PathLike[str],
    images_dir: str | os.PathLike[str],
    out_dir: str | os.PathLike[str],
    settings: InferenceSettings | None = None,
    progress: Callable[[int, int, InferenceResult | None], None] | None = None,
) -> tuple[list[InferenceResult], Path]:
    settings = settings or InferenceSettings()
    package = discover_model_package(model_path)
    image_root = Path(images_dir).expanduser().resolve()
    output_root = Path(out_dir).expanduser().resolve()
    if not image_root.is_dir():
        raise FileNotFoundError(f"Image folder not found: {image_root}")
    output_root.mkdir(parents=True, exist_ok=True)

    info = load_info_yml(package.info_yml)
    mode = settings.mode.lower()
    if mode == "auto":
        binary_thr = int(info["BinaryThresholdAuto"])
        judge_thr = float(info["JudgeThresholdAuto"])
    else:
        binary_thr = int(info["BinaryThresholdManual"])
        judge_thr = float(info["JudgeThresholdManual"])

    meta = load_meta_json(package.meta_json)
    ai_model = str(info.get("AiModel", "Unknown"))
    is_norm_model = "padim" in ai_model.lower() or "autoencoder" in ai_model.lower()
    do_normalize = settings.normalize_mode == "always" or (settings.normalize_mode == "auto" and is_norm_model)
    if do_normalize and meta is None:
        do_normalize = False

    image_paths = sorted(p for p in image_root.iterdir() if p.suffix.lower() in IMAGE_EXTENSIONS)
    if not image_paths:
        raise RuntimeError(f"No supported images found in: {image_root}")

    core = Core()
    model = core.read_model(str(package.model_xml))
    compiled = core.compile_model(model, settings.device)
    input_any = compiled.input(0)
    input_shape = list(input_any.shape)
    input_dtype = input_any.element_type.to_dtype()

    rows = []
    results: list[InferenceResult] = []
    for index, path in enumerate(image_paths, start=1):
        img = imread_unicode_safe(str(path))
        if img is None:
            continue

        x = preprocess_to_model(img, input_shape, input_dtype, settings)
        inference_result = compiled([x])
        raw_map = squeeze_to_hw(pick_anomaly_map_output(compiled, inference_result))
        norm_map = normalize_anomaly_map(raw_map, meta) if do_normalize else raw_map
        cand8 = to_uint8_map_like_sample(norm_map)
        mask01 = apply_binary_threshold(cand8, binary_thr, settings)
        mask01 = remove_small_blobs(mask01, settings.min_blob_area, settings.connectivity)
        blob_count, areas, max_blob_area = compute_blob_areas(mask01, settings.connectivity)
        sum_blob_area = int(sum(areas)) if areas else 0
        anomaly_score = compute_anomaly_score(cand8, max_blob_area, sum_blob_area, settings)
        if settings.threshold_op == ">":
            product = "NG" if anomaly_score > judge_thr else "OK"
        else:
            product = "NG" if anomaly_score >= judge_thr else "OK"

        candidate_path = binary_path = masked_path = overlay_path = ""
        if settings.save_artifacts and (not settings.save_only_ng_artifacts or product == "NG"):
            masked, overlay = build_visuals(img, cand8, mask01)
            candidate_path = _write_image(output_root / "candidate8" / f"{path.stem}_candidate.png", cand8)
            binary_path = _write_image(output_root / "binary" / f"{path.stem}_binary.png", (mask01 * 255).astype(np.uint8))
            masked_path = _write_image(output_root / "masked" / f"{path.stem}_masked.png", masked)
            overlay_path = _write_image(output_root / "overlay" / f"{path.stem}_overlay.png", overlay)

        result = InferenceResult(
            file=str(path),
            product=product,
            anomaly_score=float(anomaly_score),
            blob_count=blob_count,
            max_blob_area=max_blob_area,
            sum_blob_area=sum_blob_area,
            ng_pixels_total=int(mask01.sum()),
            candidate_path=candidate_path,
            binary_path=binary_path,
            masked_path=masked_path,
            overlay_path=overlay_path,
            width=int(cand8.shape[1]),
            height=int(cand8.shape[0]),
        )
        results.append(result)
        rows.append(
            {
                "file": result.file,
                "ai_model": ai_model,
                "normalized": do_normalize,
                "mode": settings.mode,
                "binary_threshold_255": binary_thr,
                "judge_threshold": judge_thr,
                "threshold_op": settings.threshold_op,
                "score_metric": settings.score_metric,
                "connectivity": settings.connectivity,
                "min_blob_area_filter": settings.min_blob_area,
                "anomaly_score": result.anomaly_score,
                "product": result.product,
                "blob_count": result.blob_count,
                "max_blob_area": result.max_blob_area,
                "sum_blob_area": result.sum_blob_area,
                "ng_pixels_total": result.ng_pixels_total,
                "cand8_max": int(cand8.max()),
                "cand8_mean": float(cand8.mean()),
                "raw_map_max": float(raw_map.max()),
                "raw_map_mean": float(raw_map.mean()),
                "norm_map_max": float(norm_map.max()),
                "norm_map_mean": float(norm_map.mean()),
                "candidate_path": result.candidate_path,
                "binary_path": result.binary_path,
                "masked_path": result.masked_path,
                "overlay_path": result.overlay_path,
            }
        )
        if progress:
            progress(index, len(image_paths), result)

    report_path = output_root / "results_dash_anomaly.xlsx"
    pd.DataFrame(rows).to_excel(report_path, index=False, engine="openpyxl")
    if progress:
        progress(len(image_paths), len(image_paths), None)
    return results, report_path


def export_ng_results(results: list[InferenceResult], output_path: str | os.PathLike[str]) -> Path:
    ng_results = [result for result in results if result.product == "NG"]
    if not ng_results:
        raise ValueError("There are no NG results to export.")

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NG Results"
    headers = [
        "File",
        "Score",
        "Blobs",
        "Max Blob Area",
        "Sum Blob Area",
        "NG Pixels",
        "Original Image",
        "Masked Anomaly Map",
        "Overlay",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    hyperlink_font = Font(color="0563C1", underline="single")
    for result in ng_results:
        row = [
            Path(result.file).name,
            result.anomaly_score,
            result.blob_count,
            result.max_blob_area,
            result.sum_blob_area,
            result.ng_pixels_total,
            result.file,
            result.masked_path,
            result.overlay_path,
        ]
        sheet.append(row)
        excel_row = sheet.max_row
        for column in (7, 8, 9):
            cell = sheet.cell(excel_row, column)
            if cell.value:
                cell.hyperlink = Path(str(cell.value)).resolve().as_uri()
                cell.font = hyperlink_font

    widths = [28, 12, 10, 16, 16, 12, 52, 52, 52]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    workbook.save(path)
    return path
